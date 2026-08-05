#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Verifica por EJECUCION que el Excel del CEO dice la verdad sobre WBS.md.

Uso:  .venv/bin/python 05-vista-ceo/verificar_excel.py
Sale con codigo 1 si algo FALLA. Los AVISOS no tumban la ejecucion pero se leen.

Por que existe: el 31/07/2026 el formato del WBS cambio (la ficha pasó a ir DELANTE
del estado, dentro de la misma celda) y el generador siguió leyendo solo el principio
de la celda. Resultado: 4 tareas hechas salieron como pendientes en la vista del CEO,
sin que nada avisara. Regla 25: lo que no se prueba por ejecucion, no esta verificado.

Este fichero NO comparte codigo con generar_excel.py a proposito: vuelve a leer WBS.md
con su propio analizador. Si los dos coinciden, es que el dato es bueno; si un unico
analizador se equivoca, la comprobacion no valdria nada.
"""
import json
import re
import subprocess
import sys
from pathlib import Path

RAIZ = Path(__file__).resolve().parent.parent
# Rutas por argumento SOLO para la prueba de inyeccion (regla 25: se comprueba que el
# verificador muerde metiendole un WBS roto, o -tarea 03.01.16- una fuente de REGLAS/
# LECCIONES/DECISIONES rota). En uso normal no se pasan argumentos.
WBS = Path(sys.argv[1]) if len(sys.argv) > 1 else RAIZ / "00-direccion" / "WBS.md"
XLSX = Path(sys.argv[2]) if len(sys.argv) > 2 else RAIZ / "05-vista-ceo" / "WBS_Bot_Trading_v0.9.xlsx"
CLAUDE_MD = Path(sys.argv[3]) if len(sys.argv) > 3 else RAIZ / "CLAUDE.md"
LECCIONES_MD = Path(sys.argv[4]) if len(sys.argv) > 4 else RAIZ / "00-direccion" / "LECCIONES.md"
DECISIONES_MD = Path(sys.argv[5]) if len(sys.argv) > 5 else RAIZ / "00-direccion" / "DECISIONES.md"

FALLOS, AVISOS = [], []
ESTADOS = ("pendiente", "en_curso", "hecha", "bloqueada")
ETIQUETA = {"pendiente": "Pendiente", "en_curso": "En curso", "hecha": "Hecha", "bloqueada": "Bloqueada"}


def fallo(prueba, detalle):
    FALLOS.append((prueba, detalle))
    print(f"  FALLO  {prueba}: {detalle}")


def aviso(prueba, detalle):
    AVISOS.append((prueba, detalle))
    print(f"  AVISO  {prueba}: {detalle}")


def ok(prueba, detalle=""):
    print(f"  OK     {prueba}{': ' + detalle if detalle else ''}")


# ---------------------------------------------------------------- lectura independiente
def leer_wbs():
    """Recorre el fichero entero buscando filas de tarea, sin fiarse de las secciones."""
    tareas = {}
    for linea in WBS.read_text(encoding="utf-8").splitlines():
        s = linea.strip()
        if not s.startswith("|"):
            continue
        celdas = [c.strip() for c in s.strip("|").split("|")]
        if len(celdas) < 5:
            continue
        codigo = celdas[0].strip("*` ")
        if not re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", codigo):
            continue
        if codigo in tareas:
            fallo("codigos unicos", f"{codigo} aparece dos veces en WBS.md")
        tareas[codigo] = {
            "codigo": codigo,
            "tarea": celdas[1],
            "responsable": celdas[2],
            "depende": celdas[3],
            "celda_estado": celdas[4],
        }
    return tareas


def estado_de(celda, codigo):
    """Estado declarado. Nunca lo adivina: si no hay marca, es un FALLO."""
    apertura = re.match(r"\*{0,2}(pendiente|en_curso|hecha|bloqueada)\*{0,2}\b", celda, flags=re.I)
    if apertura:
        return apertura.group(1).lower()
    negritas = re.findall(r"\*\*(pendiente|en_curso|hecha|bloqueada)\*\*", celda, flags=re.I)
    if negritas:
        if len(negritas) > 1:
            aviso("estado declarado", f"{codigo}: {len(negritas)} marcas en negrita {negritas}; vale la ultima")
        return negritas[-1].lower()
    fallo("estado declarado", f"{codigo}: su celda ESTADO no declara ningun estado -> "
                             f"el generador lo daria por 'pendiente' sin saberlo. Celda: {celda[:80]}...")
    return None


def dependencias(t, todos):
    salida = []
    for trozo in re.split(r"[,;]", t["depende"].replace("—", "").replace("–", "")):
        d = trozo.strip().strip("*` ") if not trozo.strip().endswith("*") else trozo.strip()
        if not d:
            continue
        if d.endswith("*"):
            pref = d.rstrip("*").rstrip(".")
            hijos = [c for c in todos if c.startswith(pref + ".") and c != t["codigo"]]
            if not hijos:
                fallo("dependencias existen", f"{t['codigo']} depende de '{d}' y no encaja con ninguna tarea")
            salida += hijos
        elif re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", d):
            if d not in todos:
                fallo("dependencias existen", f"{t['codigo']} depende de {d}, que no existe en el WBS")
            else:
                salida.append(d)
    return sorted(set(salida))


print(f"VERIFICACION DE LA VISTA DEL CEO\n  WBS  : {WBS}\n  Excel: {XLSX}\n")

if not XLSX.exists():
    sys.exit("FALLO: no existe el Excel. Ejecuta antes generar_excel.py")

try:
    import openpyxl
except ImportError:
    sys.exit("FALLO: falta openpyxl. Instala con: .venv/bin/pip install openpyxl")

WBS_TAREAS = leer_wbs()
for t in WBS_TAREAS.values():
    t["estado"] = estado_de(t["celda_estado"], t["codigo"])
    t["deps"] = dependencias(t, WBS_TAREAS)

wb = openpyxl.load_workbook(XLSX)
hoja = wb["TAREAS"]
EXCEL = {}
for r in range(1, hoja.max_row + 1):
    cod, est = hoja.cell(r, 1).value, hoja.cell(r, 5).value
    if isinstance(cod, str) and re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", cod):
        if cod in EXCEL:
            fallo("filas unicas", f"{cod} aparece dos veces en la hoja TAREAS")
        EXCEL[cod] = est

print("\n1. El Excel es posterior al WBS")
if XLSX.stat().st_mtime < WBS.stat().st_mtime:
    fallo("Excel al dia", "el WBS se ha tocado despues de generar el Excel: regeneralo")
else:
    ok("Excel al dia")

print("\n2. Censo de tareas: ni una perdida, ni una inventada")
faltan = sorted(set(WBS_TAREAS) - set(EXCEL))
sobran = sorted(set(EXCEL) - set(WBS_TAREAS))
if faltan:
    fallo("censo", f"en el WBS pero NO en el Excel: {faltan}")
if sobran:
    fallo("censo", f"en el Excel pero NO en el WBS: {sobran}")
if not faltan and not sobran:
    ok("censo", f"{len(WBS_TAREAS)} tareas en los dos sitios")

print("\n3. Estado de cada tarea, leido dos veces con analizadores distintos")
discrepa = []
for cod, t in sorted(WBS_TAREAS.items()):
    if t["estado"] is None:
        continue
    esperado = ETIQUETA[t["estado"]]
    if EXCEL.get(cod) != esperado:
        discrepa.append(f"{cod}: Excel dice '{EXCEL.get(cod)}' y el WBS dice '{esperado}'")
if discrepa:
    for d in discrepa:
        fallo("estados", d)
else:
    ok("estados", f"{len(WBS_TAREAS)} coinciden")

print("\n4. Ciclos de dependencias")
ciclo = None
for inicio in WBS_TAREAS:
    pila, vistos = [(inicio, [inicio])], set()
    while pila and not ciclo:
        nodo, camino = pila.pop()
        for d in WBS_TAREAS[nodo]["deps"]:
            if d == inicio:
                ciclo = " -> ".join(camino + [d])
                break
            if d not in vistos:
                vistos.add(d)
                pila.append((d, camino + [d]))
    if ciclo:
        break
if ciclo:
    fallo("ciclos", f"dependencia circular: {ciclo}")
else:
    ok("ciclos", "ninguno")

print("\n5. La cola de SIGUIENTE respeta sus propias reglas")
sig = wb["SIGUIENTE"]
grupo, cola, esperando = None, [], []
for r in range(1, sig.max_row + 1):
    a, b = sig.cell(r, 1).value, sig.cell(r, 2).value
    if isinstance(a, str) and a.startswith("SE PUEDE TRABAJAR YA"):
        grupo = "cola"
    elif isinstance(a, str) and a.startswith("ESPERANDO"):
        grupo = "esperando"
    elif isinstance(b, str) and re.fullmatch(r"\d{2}\.\d{2}\.\d{2}", b):
        (cola if grupo == "cola" else esperando).append(b)

vivas = {c for c, t in WBS_TAREAS.items() if t["estado"] != "hecha"}
if set(cola) | set(esperando) != vivas:
    fallo("cola completa", f"faltan o sobran tareas vivas: "
                           f"{sorted(vivas ^ (set(cola) | set(esperando)))}")
else:
    ok("cola completa", f"{len(vivas)} tareas vivas repartidas en cola({len(cola)}) y espera({len(esperando)})")
if set(cola) & set(esperando):
    fallo("cola sin duplicados", f"aparecen en los dos grupos: {sorted(set(cola) & set(esperando))}")
for c in cola:
    pendientes = [d for d in WBS_TAREAS[c]["deps"] if WBS_TAREAS[d]["estado"] != "hecha"]
    if pendientes:
        fallo("cola trabajable", f"{c} esta en 'se puede trabajar ya' pero espera a {pendientes}")
for c in esperando:
    pendientes = [d for d in WBS_TAREAS[c]["deps"] if WBS_TAREAS[d]["estado"] != "hecha"]
    if not pendientes:
        fallo("espera justificada", f"{c} esta en 'esperando' y no le falta ninguna dependencia")
if not FALLOS or all(p[0] not in ("cola trabajable", "espera justificada") for p in FALLOS):
    ok("orden de la cola", "cada tarea esta en el grupo que le toca")

print("\n6. Las cuentas del PANEL (formulas evaluadas de verdad)")
try:
    import formulas  # opcional: pesado, pero es la unica prueba real de las formulas
    modelo = formulas.ExcelModel().loads(str(XLSX)).finish()
    sol = modelo.calculate()
    def celda(ref):
        v = sol.get(f"'[{XLSX.name}]PANEL'!{ref}")
        try:
            return v.value[0, 0]
        except Exception:
            return v
    total = celda("B11")
    hechas = celda("C11")
    hechas_wbs = sum(1 for t in WBS_TAREAS.values() if t["estado"] == "hecha")
    if int(total) != len(WBS_TAREAS):
        fallo("panel", f"el panel suma {total} tareas y el WBS tiene {len(WBS_TAREAS)}")
    elif int(hechas) != hechas_wbs:
        fallo("panel", f"el panel cuenta {hechas} hechas y el WBS tiene {hechas_wbs}")
    else:
        ok("panel", f"total {int(total)} · hechas {int(hechas)}, cuadra con el WBS")
except ImportError:
    aviso("panel", "sin el paquete 'formulas' no se pueden evaluar los COUNTIFS: NO VERIFICADO")

print("\n7. Las tareas hechas tienen su entregable en disco")
sin_artefacto = []
for cod, t in sorted(WBS_TAREAS.items()):
    if t["estado"] != "hecha":
        continue
    corte = re.search(r"\*\*hecha\*\*|^hecha", t["celda_estado"], flags=re.I)
    cola_celda = t["celda_estado"][corte.end():] if corte else t["celda_estado"]
    for nombre in set(re.findall(r"`([\w./-]+\.(?:md|py|json|csv|txt|xlsx))`", cola_celda)):
        base = Path(nombre).name
        if not list(RAIZ.rglob(base)):
            sin_artefacto.append(f"{cod} dice haber entregado {nombre} y no esta en el repositorio")
if sin_artefacto:
    for s in sin_artefacto:
        aviso("entregables", s)
else:
    ok("entregables", "todos los ficheros citados por las tareas hechas existen")

print("\n8. Regla 24: los datos siguen fuera de git")
try:
    salida = subprocess.run(["git", "status", "--porcelain", "02-datos/"], cwd=RAIZ,
                            capture_output=True, text=True, timeout=30).stdout.strip()
    if salida:
        fallo("regla 24", f"git ve ficheros dentro de 02-datos/:\n{salida[:400]}")
    else:
        ok("regla 24", "git no ve nada dentro de 02-datos/")
except Exception as e:  # noqa: BLE001
    aviso("regla 24", f"no se pudo comprobar: {e}")

# ---------------------------------------------------------------- tarea 03.01.16
# Las hojas REGLAS, LECCIONES y DECISIONES ya no salen de la tabla espejo del WBS: salen de
# CLAUDE.md, 00-direccion/LECCIONES.md y 00-direccion/DECISIONES.md. Este verificador vuelve a
# contar esas tres fuentes con SU PROPIO analizador (no importa nada de generar_excel.py, por el
# mismo motivo que el resto del fichero: si los dos coinciden es que el dato es bueno) y compara
# contra el recuento que trae la hoja. D-25: la comparacion es SIEMPRE contra el recuento vivo,
# nunca contra un numero escrito a mano aqui.


def cuerpo_tras_titulo(texto, fragmento_titulo):
    """Cuerpo de la primera sección '## ...' cuyo título contiene el fragmento (sin
    mayúsculas), hasta la siguiente '## '. None si no se encuentra."""
    lineas = texto.splitlines()
    inicio = next((i for i, l in enumerate(lineas)
                   if l.startswith("## ") and fragmento_titulo.lower() in l.lower()), None)
    if inicio is None:
        return None
    fin = next((j for j in range(inicio + 1, len(lineas)) if lineas[j].startswith("## ")), len(lineas))
    return "\n".join(lineas[inicio + 1:fin])


def contar_columna_a(nombre_hoja, patron):
    """Cuenta, en la columna A de una hoja, cuantas celdas cumplen un patron completo.
    None si la hoja no existe (para poder distinguir 'hoja ausente' de 'hoja vacia')."""
    if nombre_hoja not in wb.sheetnames:
        return None
    hoja = wb[nombre_hoja]
    total = 0
    for fila in hoja.iter_rows(min_col=1, max_col=1):
        v = fila[0].value
        if v is None:
            continue
        if re.fullmatch(patron, str(v).strip()):
            total += 1
    return total


print("\n9. Las hojas REGLAS, LECCIONES y DECISIONES traen el recuento VIVO de su fuente")
cuerpo_reglas = cuerpo_tras_titulo(CLAUDE_MD.read_text(encoding="utf-8"), "reglas")
if cuerpo_reglas is None:
    fallo("hoja REGLAS", "CLAUDE.md no tiene ninguna sección '## Las N reglas' que contar")
else:
    # los sub-puntos indentados de la regla 9 ('   1. Prueba ejecutada...') no cuentan: el
    # numero tiene que abrir la linea, sin permitir que la indentacion se cuele (ver L-018).
    reglas_fuente = len(re.findall(r"^\d+\. ", cuerpo_reglas, flags=re.M))
    reglas_excel = contar_columna_a("REGLAS", r"\d+")
    if reglas_excel is None:
        fallo("hoja REGLAS", "la hoja 'REGLAS' no existe en el Excel")
    elif reglas_excel != reglas_fuente:
        fallo("hoja REGLAS", f"el Excel trae {reglas_excel} reglas y CLAUDE.md tiene {reglas_fuente} "
                             f"ahora mismo (recuento vivo, no el de esta ficha)")
    else:
        ok("hoja REGLAS", f"{reglas_fuente} reglas, igual en el Excel y en CLAUDE.md")

texto_lecciones = LECCIONES_MD.read_text(encoding="utf-8")
lecciones_fuente = len(re.findall(r"^## L-\d+ · ", texto_lecciones, flags=re.M))
lecciones_excel = contar_columna_a("LECCIONES", r"L-\d+")
if lecciones_excel is None:
    fallo("hoja LECCIONES", "la hoja 'LECCIONES' no existe en el Excel")
elif lecciones_excel != lecciones_fuente:
    fallo("hoja LECCIONES", f"el Excel trae {lecciones_excel} lecciones y "
                            f"00-direccion/LECCIONES.md tiene {lecciones_fuente} ahora mismo "
                            f"(recuento vivo, no el de esta ficha)")
else:
    ok("hoja LECCIONES", f"{lecciones_fuente} lecciones, igual en el Excel y en LECCIONES.md")

texto_decisiones = DECISIONES_MD.read_text(encoding="utf-8")
decisiones_fuente = len(re.findall(r"^## D-\d+ · \d{4}-\d{2}-\d{2} · ", texto_decisiones, flags=re.M))
decisiones_excel = contar_columna_a("DECISIONES", r"D-\d+")
if decisiones_excel is None:
    fallo("hoja DECISIONES", "la hoja 'DECISIONES' no existe en el Excel")
elif decisiones_excel != decisiones_fuente:
    fallo("hoja DECISIONES", f"el Excel trae {decisiones_excel} decisiones y "
                             f"00-direccion/DECISIONES.md tiene {decisiones_fuente} ahora mismo "
                             f"(recuento vivo, no el de esta ficha)")
else:
    ok("hoja DECISIONES", f"{decisiones_fuente} decisiones, igual en el Excel y en DECISIONES.md")

# ---------------------------------------------------------------- posicion del estado
# Hoy (antes de esta prueba) solo se contaba CUANTAS marcas de estado en negrita hay por
# celda (aviso, no fallo); nunca DONDE caen. Una marca sobrante en mitad de la celda no
# falsea el ESTADO (partir_estado usa la ULTIMA), pero descoloca lo que separa FICHA de
# RESULTADO: la ficha pierde su cierre y el resultado arrastra texto que no le corresponde.
#
# Una marca de estado en negrita esta BIEN colocada si, y solo si, pasa una de las tres:
#   (a) CIERRA una transicion real: la preceden '—'/'-'/'–' (el patron que usa todo el WBS:
#       '... — **hecha** fecha — resultado').
#   (b) ABRE la celda (formato A: 'hecha 31/07 — resultado'), sin nada delante salvo blancos.
#   (c) es una CITA de texto entre comillas invertidas ('`**hecha**`'), que no declara nada:
#       cita el patron como ejemplo, como hace el propio WBS al describir este mismo bug.
# Cualquier otra marca no cierra nada ni se cita: esta EN MEDIO, y es la prueba nueva.
# Verificado contra el WBS.md real (61 celdas de tarea, 42 marcas): CERO falsos positivos
# antes de inyectar nada — la unica marca que no encajaba en (a)/(b)/(c) era un defecto propio
# de esta reescritura ('cerró **hecha** el 03/08'), reparado antes de escribir esta prueba.
def marca_bien_colocada(celda, ini, fin):
    antes = celda[:ini]
    antes3 = celda[max(0, ini - 2):ini]
    despues1 = celda[fin:fin + 1]
    cierra_transicion = antes3.rstrip().endswith(("—", "-", "–"))
    abre_celda = antes.strip() == ""
    es_cita = antes3.endswith("`") and despues1 == "`"
    return cierra_transicion or abre_celda or es_cita


print("\n10. Posición del estado: ninguna marca de estado en negrita 'en medio' de la celda")
descolocadas = []
marcas_revisadas = 0
for cod, t in sorted(WBS_TAREAS.items()):
    celda = t["celda_estado"]
    for m in re.finditer(r"\*\*(pendiente|en_curso|hecha|bloqueada)\*\*", celda, flags=re.I):
        marcas_revisadas += 1
        if not marca_bien_colocada(celda, m.start(), m.end()):
            entorno = celda[max(0, m.start() - 25):m.end() + 15].replace("\n", " ")
            descolocadas.append(f"{cod}: la marca '{m.group(0)}' (posición {m.start()}) no cierra "
                                f"una transición, no abre la celda ni va citada entre comillas "
                                f"invertidas -> está EN MEDIO y descoloca FICHA/RESULTADO. "
                                f"Contexto: ...{entorno}...")
if descolocadas:
    for d in descolocadas:
        fallo("posición del estado", d)
else:
    ok("posición del estado", f"{marcas_revisadas} marcas revisadas en {len(WBS_TAREAS)} celdas, ninguna en medio")

print("\n" + "=" * 70)
if FALLOS:
    print(f"RESULTADO: {len(FALLOS)} FALLOS y {len(AVISOS)} avisos. El Excel NO es de fiar.")
    for p, d in FALLOS:
        print(f"  - [{p}] {d}")
    sys.exit(1)
print(f"RESULTADO: sin fallos. {len(AVISOS)} avisos." if AVISOS else "RESULTADO: sin fallos ni avisos.")
for p, d in AVISOS:
    print(f"  - [{p}] {d}")
sys.exit(0)
