#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Vista Excel del CEO — se genera LEYENDO 00-direccion/WBS.md, nunca al reves.

Uso:  .venv/bin/python 05-vista-ceo/generar_excel.py
Salida: 05-vista-ceo/WBS_Bot_Trading_v0.9.xlsx

Regla 11 del WBS: el texto se actualiza siempre; este Excel se regenera para la
revision de los lunes y para las puertas. Como lee el markdown, no puede quedarse
desfasado respecto a la fuente de verdad.
"""
import json
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("FALTA openpyxl. Crea el entorno:  python3 -m venv .venv && .venv/bin/pip install openpyxl formulas\n"
             "y ejecuta este script con .venv/bin/python (el Python del sistema no deja instalar paquetes).")
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

RAIZ = Path(__file__).resolve().parent.parent
WBS = RAIZ / "00-direccion" / "WBS.md"
SALIDA = Path(__file__).resolve().parent / "WBS_Bot_Trading_v0.9.xlsx"

# Tarea 03.01.16: REGLAS, LECCIONES y DECISIONES se leen de sus fuentes primarias,
# nunca de las tablas espejo del WBS (que quedaron desfasadas: D-16 vacio la de reglas,
# y las de lecciones/decisiones se quedaron congeladas en versiones antiguas).
CLAUDE_MD = RAIZ / "CLAUDE.md"
LECCIONES_MD = RAIZ / "00-direccion" / "LECCIONES.md"
DECISIONES_MD = RAIZ / "00-direccion" / "DECISIONES.md"

# ---------------------------------------------------------------- estilo
ARIAL = "Arial"
F_TITULO = Font(name=ARIAL, size=13, bold=True, color="FFFFFF")
F_HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
F_FASE = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
F_TXT = Font(name=ARIAL, size=10)
F_BOLD = Font(name=ARIAL, size=10, bold=True)
F_MINI = Font(name=ARIAL, size=9, italic=True)
FILL_TITULO = PatternFill("solid", fgColor="1F3864")
FILL_HDR = PatternFill("solid", fgColor="2F5496")
FILL_FASE = PatternFill("solid", fgColor="4472C4")
FILL_NOTA = PatternFill("solid", fgColor="FFF2CC")
FILL_HECHA = PatternFill("solid", fgColor="C6EFCE")
FILL_CURSO = PatternFill("solid", fgColor="FFEB9C")
FILL_BLOQ = PatternFill("solid", fgColor="FFC7CE")
FILL_CEO = PatternFill("solid", fgColor="FCE4D6")
BORDE = Border(*[Side(style="thin", color="B0B0B0")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTRO = Alignment(horizontal="center", vertical="center", wrap_text=True)

ESTADOS = ["pendiente", "en_curso", "hecha", "bloqueada"]
ETIQUETA = {"pendiente": "Pendiente", "en_curso": "En curso", "hecha": "Hecha", "bloqueada": "Bloqueada"}


# ---------------------------------------------------------------- lectura del markdown
def leer_secciones(texto):
    """Devuelve {titulo_de_seccion: cuerpo} partiendo por encabezados '## '."""
    secciones, titulo, buffer = {}, "__inicio__", []
    for linea in texto.splitlines():
        if linea.startswith("## "):
            secciones[titulo] = "\n".join(buffer)
            titulo, buffer = linea[3:].strip(), []
        else:
            buffer.append(linea)
    secciones[titulo] = "\n".join(buffer)
    return secciones


def tablas(cuerpo):
    """Extrae todas las tablas markdown de un cuerpo. Cada tabla: lista de listas."""
    resultado, actual = [], []
    for linea in cuerpo.splitlines():
        s = linea.strip()
        if s.startswith("|") and s.endswith("|"):
            celdas = [c.strip() for c in s.strip("|").split("|")]
            if all(re.fullmatch(r":?-{2,}:?", c) for c in celdas):
                continue  # separador |---|---|
            actual.append(celdas)
        elif actual:
            resultado.append(actual)
            actual = []
    if actual:
        resultado.append(actual)
    return resultado


def limpiar(texto):
    """Quita negritas, cursivas y backticks de markdown; deja el texto plano."""
    t = re.sub(r"\*\*(.+?)\*\*", r"\1", texto)
    t = re.sub(r"(?<!\w)\*(.+?)\*(?!\w)", r"\1", t)
    t = t.replace("`", "")
    return t.strip()


AVISOS = []


def partir_estado(celda, codigo=""):
    """Saca (estado, ficha, resultado) de la celda ESTADO del WBS.

    Convive con los dos formatos que usa WBS.md:
      A) 'hecha 31/07 — artefacto'            -> el estado abre la celda
      B) 'FICHA (31/07): ... — **hecha** 31/07 — artefacto'
         -> la ficha va delante y el estado aparece EN NEGRITA en medio.
    En el formato B el estado nunca se busca en texto llano, porque frases como
    "sin esto sigue bloqueada" o "decision del CEO pendiente" darian un falso
    positivo: solo cuenta la palabra en negrita.
    """
    bruto = celda.strip()
    plano = limpiar(bruto)

    # formato A: la celda abre con el estado (con o sin negrita)
    apertura = re.match(r"\*{0,2}(pendiente|en_curso|hecha|bloqueada)\*{0,2}\b", bruto, flags=re.I)
    if apertura:
        estado = apertura.group(1).lower()
        resto = limpiar(bruto[apertura.end():]).lstrip(" —-–:")
        return estado, "", resto

    # formato B: ultima marca de estado EN NEGRITA
    marcas = list(re.finditer(r"\*\*(pendiente|en_curso|hecha|bloqueada)\*\*", bruto, flags=re.I))
    if marcas:
        m = marcas[-1]
        estado = m.group(1).lower()
        ficha = limpiar(bruto[:m.start()]).rstrip(" —-–:")
        resultado = limpiar(bruto[m.end():]).lstrip(" —-–:")
        if len(marcas) > 1:
            AVISOS.append(f"{codigo}: {len(marcas)} marcas de estado en negrita; me quedo con la ultima ({estado})")
        return estado, ficha, resultado

    # sin marca: no se supone nada, se avisa
    if plano:
        AVISOS.append(f"{codigo}: la celda ESTADO no declara estado ({plano[:60]}...) -> lo doy por pendiente")
    return "pendiente", "", plano


def recortar(texto, tope=700, fuente="00-direccion/WBS.md"):
    """La vista del CEO no reproduce fichas de 3.000 caracteres: apunta a la fuente."""
    if len(texto) <= tope:
        return texto
    corte = texto.rfind(" ", 0, tope)
    return texto[:corte if corte > 0 else tope] + f" […] texto completo en {fuente}"


def bloques(texto, patron_cabecera):
    """Parte un texto en bloques a partir de una cabecera con grupos de captura.
    Devuelve una lista de (grupos_de_la_cabecera, cuerpo_hasta_la_siguiente_cabecera)."""
    cabeceras = list(re.finditer(patron_cabecera, texto, flags=re.M))
    salida = []
    for i, m in enumerate(cabeceras):
        fin = cabeceras[i + 1].start() if i + 1 < len(cabeceras) else len(texto)
        salida.append((m.groups(), texto[m.end():fin]))
    return salida


def campo(cuerpo, etiqueta):
    """Extrae el texto de un campo '**Etiqueta:** ...' hasta el siguiente campo en negrita
    ('**Otra cosa:**') o el final del bloque. Devuelve '' si la etiqueta no aparece."""
    m = re.search(r"\*\*" + re.escape(etiqueta) + r":\*\*(.*?)(?=\n\*\*[^\n]{1,80}:\*\*|\Z)", cuerpo, flags=re.S)
    return limpiar(m.group(1)) if m else ""


def es_codigo(celda):
    return bool(re.fullmatch(r"\d{2}(\.\d{2}){0,2}", limpiar(celda)))


TEXTO = WBS.read_text(encoding="utf-8")
SEC = leer_secciones(TEXTO)

# ---------------------------------------------------------------- fuentes primarias
# Tarea 03.01.16: REGLAS/LECCIONES/DECISIONES se construyen aqui, leyendo CLAUDE.md,
# LECCIONES.md y DECISIONES.md directamente. Nunca de las tablas espejo del WBS.

# ---- REGLAS: CLAUDE.md, seccion "## Las N reglas". Los sub-puntos indentados de la
# regla 9 ("   1. Prueba ejecutada...") NO cuentan: por eso se exige que el numero
# vaya al principio absoluto de la linea, sin permitir que un strip() los cuele.
TEXTO_CLAUDE = CLAUDE_MD.read_text(encoding="utf-8")
SEC_CLAUDE = leer_secciones(TEXTO_CLAUDE)
CUERPO_REGLAS = next((c for t, c in SEC_CLAUDE.items() if t.lower().startswith("las ") and "reglas" in t.lower()), "")
REGLAS = []
for linea in CUERPO_REGLAS.splitlines():
    m = re.match(r"(\d+)\. (.+)", linea)
    if m:
        REGLAS.append((int(m.group(1)), limpiar(m.group(2))))

# ---- LECCIONES: 00-direccion/LECCIONES.md, una fila por cabecera "## L-NNN · ...".
TEXTO_LECCIONES = LECCIONES_MD.read_text(encoding="utf-8")
LECCIONES = []
for (id_lec, sintoma), cuerpo_lec in bloques(TEXTO_LECCIONES, r"^## (L-\d+) · (.+)$"):
    LECCIONES.append({
        "id": id_lec,
        "sintoma": limpiar(sintoma),
        "causa": campo(cuerpo_lec, "Causa raiz"),
        "regla": campo(cuerpo_lec, "Regla"),
        "evento": campo(cuerpo_lec, "Evento"),
    })

# ---- DECISIONES: 00-direccion/DECISIONES.md, una fila por cabecera "## D-N · fecha · ...".
TEXTO_DECISIONES = DECISIONES_MD.read_text(encoding="utf-8")
DECISIONES = []
for (id_dec, fecha_dec, titulo_dec), cuerpo_dec in bloques(TEXTO_DECISIONES, r"^## (D-\d+) · (\d{4}-\d{2}-\d{2}) · (.+)$"):
    DECISIONES.append({
        "id": id_dec,
        "fecha": fecha_dec,
        "decision": limpiar(titulo_dec),
        "detalle": limpiar(cuerpo_dec),
    })

if len(REGLAS) == 0:
    AVISOS.append("REGLAS: CLAUDE.md no dio ninguna regla numerada; revisa el encabezado '## Las N reglas'")
if len(LECCIONES) == 0:
    AVISOS.append("LECCIONES: 00-direccion/LECCIONES.md no dio ninguna entrada 'L-NNN'")
if len(DECISIONES) == 0:
    AVISOS.append("DECISIONES: 00-direccion/DECISIONES.md no dio ninguna entrada 'D-N'")


def seccion(fragmento):
    """Busca una seccion por fragmento del titulo (sin acentos ni mayusculas)."""
    for titulo, cuerpo in SEC.items():
        if fragmento.lower() in titulo.lower():
            return cuerpo
    raise KeyError(f"No encuentro la seccion '{fragmento}' en WBS.md")


# ------- tareas: recorre todas las secciones 'Fase NN — ...'
FASES = []  # (codigo, nombre, [tareas])
for titulo, cuerpo in SEC.items():
    m = re.match(r"Fase (\d{2}) — (.+)", titulo)
    if not m:
        continue
    tareas = []
    # sub-encabezados '### 04.01 Broker y cajones de datos' -> etiqueta del paquete
    paquetes = {}
    for linea in cuerpo.splitlines():
        sub = re.match(r"###\s+(\d{2}\.\d{2})\s+(.+)", linea.strip())
        if sub:
            paquetes[sub.group(1)] = limpiar(sub.group(2))
    for tabla in tablas(cuerpo):
        for fila in tabla:
            if len(fila) < 5 or not es_codigo(fila[0]):
                continue
            codigo = limpiar(fila[0])
            estado, ficha, nota = partir_estado(fila[4], codigo)
            tareas.append({
                "codigo": codigo,
                "paquete": paquetes.get(codigo[:5], ""),
                "ficha": ficha,
                "tarea": limpiar(fila[1]),
                "responsable": limpiar(fila[2]),
                "depende": limpiar(fila[3]),
                "estado": estado,
                "nota": nota,
            })
    FASES.append((m.group(1), limpiar(m.group(2)), tareas))
FASES.sort(key=lambda f: f[0])

if not FASES or sum(len(f[2]) for f in FASES) == 0:
    sys.exit("ERROR: no he leido ninguna tarea de WBS.md. Revisa el formato de las tablas.")

# ---------------------------------------------------------------- orden de trabajo
# Reglas aplicadas, todas escritas en WBS.md:
#   regla 4  -> se va en orden salvo lo marcado paralelo
#   regla 12 -> el producto va antes que el motor (motor = fases 03 y 07)
# Una tarea se puede trabajar cuando TODAS sus dependencias estan hechas.
TODAS = [t for _, _, tareas in FASES for t in tareas]
POR_CODIGO = {t["codigo"]: t for t in TODAS}


def dependencias(t):
    """Codigos concretos de los que depende t. Resuelve comodines tipo '02.02.*'."""
    bruto = t["depende"].replace("—", "").replace("–", "")
    salida = []
    for trozo in re.split(r"[,;]", bruto):
        d = trozo.strip()
        if not d:
            continue
        if d.endswith("*"):
            prefijo = d.rstrip("*").rstrip(".")
            salida += [c for c in POR_CODIGO if c.startswith(prefijo + ".") and c != t["codigo"]]
        elif d in POR_CODIGO:
            salida.append(d)
    return sorted(set(salida))


def es_motor(t):
    """Carril de motor: fases 03 y 07, mas lo que su propio texto marque como tal."""
    return t["codigo"][:2] in ("03", "07") or "carril de motor" in t["tarea"].lower()


for t in TODAS:
    t["deps"] = dependencias(t)
    t["falta"] = [d for d in t["deps"] if POR_CODIGO[d]["estado"] != "hecha"]
    t["lista"] = not t["falta"]
    t["carril"] = "MOTOR" if es_motor(t) else "PRODUCTO"
    t["es_ceo"] = "CEO" in t["responsable"]
# DESBLOQUEA: tareas vivas que dependen de ella (directas) y en cadena (transitivas).
HIJOS = {t["codigo"]: [o["codigo"] for o in TODAS if t["codigo"] in o["deps"] and o["estado"] != "hecha"] for t in TODAS}
for t in TODAS:
    t["desbloquea"] = len(HIJOS[t["codigo"]])
    vistos, pila = set(), list(HIJOS[t["codigo"]])
    while pila:
        c = pila.pop()
        if c in vistos or c == t["codigo"]:
            continue
        vistos.add(c)
        pila += HIJOS[c]
    t["cadena"] = len(vistos)

# Cola de trabajo: primero lo empezado, luego producto antes que motor, y a igualdad,
# orden de codigo WBS.
ORDEN_ESTADO = {"en_curso": 0, "bloqueada": 1, "pendiente": 1}
COLA = sorted(
    (t for t in TODAS if t["estado"] != "hecha" and t["lista"]),
    key=lambda t: (ORDEN_ESTADO[t["estado"]], 0 if t["carril"] == "PRODUCTO" else 1, t["codigo"]),
)
ESPERANDO = sorted(
    (t for t in TODAS if t["estado"] != "hecha" and not t["lista"]),
    key=lambda t: t["codigo"],
)

wb = openpyxl.Workbook()


def cabecera(hoja, titulo, ancho, subtitulo=None):
    hoja["A1"] = titulo
    hoja["A1"].font = F_TITULO
    hoja["A1"].alignment = Alignment(vertical="center")
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ancho)
    for c in range(1, ancho + 1):
        hoja.cell(row=1, column=c).fill = FILL_TITULO
    hoja.row_dimensions[1].height = 24
    if subtitulo:
        hoja["A2"] = subtitulo
        hoja["A2"].font = F_MINI
        hoja["A2"].alignment = WRAP
        hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ancho)
        for c in range(1, ancho + 1):
            hoja.cell(row=2, column=c).fill = FILL_NOTA
        return 4
    return 3


def fila_hdr(hoja, fila, titulos):
    for j, h in enumerate(titulos, start=1):
        c = hoja.cell(row=fila, column=j, value=h)
        c.font = F_HDR
        c.fill = FILL_HDR
        c.alignment = CENTRO
        c.border = BORDE
    hoja.row_dimensions[fila].height = 28


def anchos(hoja, valores):
    for j, w in enumerate(valores, start=1):
        hoja.column_dimensions[get_column_letter(j)].width = w


def volcar(hoja, fila, filas, rellenos=None):
    """Escribe filas de texto con borde y wrap. Devuelve la ultima fila usada."""
    for datos in filas:
        fila += 1
        for j, v in enumerate(datos, start=1):
            c = hoja.cell(row=fila, column=j, value=v)
            c.border = BORDE
            c.font = F_TXT
            c.alignment = WRAP
        if rellenos:
            for j, f in rellenos:
                hoja.cell(row=fila, column=j).fill = f
    return fila


# ================================================================ 1. TAREAS
tar = wb.active
tar.title = "TAREAS"
f = cabecera(
    tar,
    "TODAS LAS TAREAS — WBS v0.9 · generado de 00-direccion/WBS.md",
    7,
    "Una fila por tarea. Filtra por ESTADO con la flecha de la cabecera. Verde = hecha · amarillo = en curso · rojo = bloqueada. "
    "Las tareas del CEO llevan la columna RESPONSABLE en naranja. Los textos largos se cortan y se apunta al WBS: esto es la vista, no la fuente.",
)
fila_hdr(tar, f, ["CODIGO", "QUE SE HACE", "RESPONSABLE", "DEPENDE DE", "ESTADO", "RESULTADO / DONDE ESTA", "FICHA (LO QUE SE ENCARGO)"])
hdr_tareas = f
r = f
primera_tarea = f + 1
for cod, nombre, tareas in FASES:
    r += 1
    c = tar.cell(row=r, column=1, value=cod)
    c.number_format = "@"
    tar.cell(row=r, column=2, value=nombre.upper())
    for j in range(1, 8):
        cc = tar.cell(row=r, column=j)
        cc.fill = FILL_FASE
        cc.font = F_FASE
        cc.border = BORDE
        cc.alignment = WRAP
    tar.row_dimensions[r].height = 20
    paquete_actual = None
    for t in tareas:
        if t["paquete"] and t["paquete"] != paquete_actual:
            paquete_actual = t["paquete"]
            r += 1
            c = tar.cell(row=r, column=1, value=t["codigo"][:5])
            c.number_format = "@"
            tar.cell(row=r, column=2, value=paquete_actual)
            for j in range(1, 8):
                cc = tar.cell(row=r, column=j)
                cc.fill = FILL_NOTA
                cc.font = F_BOLD
                cc.border = BORDE
                cc.alignment = WRAP
        r += 1
        valores = [t["codigo"], t["tarea"], t["responsable"], t["depende"], ETIQUETA[t["estado"]],
                   recortar(t["nota"]), recortar(t["ficha"], 500)]
        for j, v in enumerate(valores, start=1):
            c = tar.cell(row=r, column=j, value=v)
            c.border = BORDE
            c.font = F_TXT
            c.alignment = WRAP
            if j == 1:
                c.number_format = "@"
        if "CEO" in t["responsable"]:
            tar.cell(row=r, column=3).fill = FILL_CEO
ultima_tarea = r

rango = f"E{primera_tarea}:E{ultima_tarea}"
dv = DataValidation(type="list", formula1='"Pendiente,En curso,Hecha,Bloqueada"', allow_blank=True)
tar.add_data_validation(dv)
dv.add(rango)
for etiqueta, relleno in [("Hecha", FILL_HECHA), ("En curso", FILL_CURSO), ("Bloqueada", FILL_BLOQ)]:
    tar.conditional_formatting.add(rango, CellIsRule(operator="equal", formula=[f'"{etiqueta}"'], fill=relleno))
tar.auto_filter.ref = f"A{hdr_tareas}:G{ultima_tarea}"
tar.freeze_panes = f"A{primera_tarea}"
anchos(tar, [10, 58, 19, 16, 12, 56, 50])

# ================================================================ 2. PANEL (primero)
pan = wb.create_sheet("PANEL", 0)
f = cabecera(pan, "PANEL — BOT DE TRADING · puerta del mes (GM): 1 de septiembre de 2026", 7)
fila_hdr(pan, f, ["FASE", "TAREAS", "HECHAS", "EN CURSO", "PENDIENTES", "BLOQUEADAS", "% AVANCE"])
r = f
COD = f"TAREAS!$A${primera_tarea}:$A${ultima_tarea}"
EST = f"TAREAS!$E${primera_tarea}:$E${ultima_tarea}"
primera_fase = f + 1
for cod, nombre, tareas in FASES:
    r += 1
    # patron ??.??.?? : cuenta solo codigos de tarea completos, nunca las filas
    # de fase ('04') ni las de paquete ('04.01')
    patron = f"{cod}.??.??"
    corto = re.split(r" — | \(", nombre)[0]
    pan.cell(row=r, column=1, value=f"{cod}. {corto.capitalize()}")
    pan.cell(row=r, column=2, value=f'=COUNTIFS({COD},"{patron}")')
    for j, et in enumerate(["Hecha", "En curso", "Pendiente", "Bloqueada"], start=3):
        pan.cell(row=r, column=j, value=f'=COUNTIFS({COD},"{patron}",{EST},"{et}")')
    pan.cell(row=r, column=7, value=f"=IFERROR(C{r}/B{r},0)").number_format = "0%"
    for j in range(1, 8):
        cc = pan.cell(row=r, column=j)
        cc.border = BORDE
        cc.font = F_TXT
        cc.alignment = WRAP if j == 1 else CENTRO
total = r + 1
pan.cell(row=total, column=1, value="TOTAL PROYECTO").font = F_BOLD
for j, col in zip(range(2, 7), "BCDEF"):
    cc = pan.cell(row=total, column=j, value=f"=SUM({col}{primera_fase}:{col}{total-1})")
    cc.font = F_BOLD
    cc.alignment = CENTRO
cc = pan.cell(row=total, column=7, value=f"=IFERROR(C{total}/B{total},0)")
cc.number_format = "0%"
cc.font = F_BOLD
cc.alignment = CENTRO
for j in range(1, 8):
    pan.cell(row=total, column=j).border = BORDE
    pan.cell(row=total, column=j).fill = FILL_NOTA
pan.conditional_formatting.add(
    f"G{primera_fase}:G{total}",
    DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color="4472C4", showValue=True),
)

# lo que esta vivo ahora, leido del propio WBS
vivas = [(t["codigo"], t["tarea"], t["estado"]) for _, _, ts in FASES for t in ts if t["estado"] in ("en_curso", "bloqueada")]
r = total + 2
pan.cell(row=r, column=1, value="ESTADO DE HOY (leido de WBS.md)").font = Font(name=ARIAL, size=11, bold=True)
r += 1
fila_hdr(pan, r, ["CODIGO", "TAREA", "", "", "", "", "ESTADO"])
pan.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
for codigo, nombre, estado in vivas:
    r += 1
    pan.cell(row=r, column=1, value=codigo).number_format = "@"
    pan.cell(row=r, column=2, value=nombre)
    pan.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    pan.cell(row=r, column=7, value=ETIQUETA[estado])
    for j in range(1, 8):
        cc = pan.cell(row=r, column=j)
        cc.border = BORDE
        cc.font = F_TXT
        cc.alignment = WRAP
    pan.cell(row=r, column=7).fill = FILL_CURSO if estado == "en_curso" else FILL_BLOQ
    pan.cell(row=r, column=7).alignment = CENTRO

r += 2
for etiqueta, texto in [
    ("PROXIMA PUERTA", "G1 — elegir mercado y tamaño de vela (criterios en la hoja PUERTAS)"),
    ("OBJETIVO DEL MES", "Llegar al 1 de septiembre con respuesta a: ¿existe alguna estrategia que merezca pasar a demo?"),
    ("FUENTE DE VERDAD", "00-direccion/WBS.md — este Excel se genera de ahi. Si discrepan, manda el texto."),
]:
    pan.cell(row=r, column=1, value=etiqueta).font = F_BOLD
    pan.cell(row=r, column=2, value=texto).font = F_TXT
    pan.cell(row=r, column=2).alignment = WRAP
    pan.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    r += 1
r += 1
r += 2
pan.cell(row=r, column=1, value="LEYENDA").font = F_BOLD
for i, (txt, relleno) in enumerate([("Hecha", FILL_HECHA), ("En curso", FILL_CURSO), ("Pendiente", None), ("Bloqueada", FILL_BLOQ)]):
    c = pan.cell(row=r + 1 + i, column=1, value=txt)
    c.font = F_TXT
    c.border = BORDE
    if relleno:
        c.fill = relleno
anchos(pan, [36, 10, 10, 11, 12, 12, 14])

# ================================================================ 2 bis. SIGUIENTE
sig = wb.create_sheet("SIGUIENTE", 1)
f = cabecera(
    sig,
    "QUE SE HACE AHORA Y EN QUE ORDEN",
    9,
    "Orden calculado de WBS.md: (1) una tarea entra en la cola cuando TODAS sus dependencias estan hechas; "
    "(2) primero lo ya empezado; (3) despues el PRODUCTO antes que el MOTOR (regla 12: el motor no manda); "
    "(4) a igualdad, orden de codigo WBS (regla 4). MOTOR = fases 03 y 07 y lo que su texto marque como tal. "
    "DESBLOQUEA = tareas vivas que la esperan directamente · EN CADENA = todas las que quedan detras contando el efecto domino.",
)

siguiente_equipo = next((t for t in COLA if not t["es_ceo"]), None)
siguiente_ceo = next((t for t in COLA if t["es_ceo"]), None)
r = f
for etiqueta, t, relleno in [
    ("AHORA — EQUIPO", siguiente_equipo, FILL_CURSO),
    ("AHORA — CEO", siguiente_ceo, FILL_CEO),
]:
    sig.cell(row=r, column=1, value=etiqueta).font = Font(name=ARIAL, size=11, bold=True)
    sig.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    texto = f"{t['codigo']} — {t['tarea']}" if t else "nada disponible: todo lo pendiente espera a otra tarea"
    quien = t["responsable"] if t else "—"
    sig.cell(row=r, column=3, value=texto).font = F_TXT
    sig.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    sig.cell(row=r, column=9, value=quien).font = F_TXT
    for j in range(1, 10):
        cc = sig.cell(row=r, column=j)
        cc.fill = relleno
        cc.border = BORDE
        cc.alignment = WRAP
    sig.row_dimensions[r].height = 34
    r += 1

cuello = max(COLA, key=lambda t: t["cadena"]) if COLA else None
if cuello:
    sig.cell(row=r, column=1, value="CUELLO DE BOTELLA").font = Font(name=ARIAL, size=11, bold=True)
    sig.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    sig.cell(row=r, column=3, value=(
        f"{cuello['codigo']} — {cuello['tarea']}. Detras de ella esperan {cuello['cadena']} de las "
        f"{sum(1 for t in TODAS if t['estado'] != 'hecha')} tareas vivas."
    )).font = F_TXT
    sig.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    sig.cell(row=r, column=9, value=cuello["responsable"]).font = F_TXT
    for j in range(1, 10):
        cc = sig.cell(row=r, column=j)
        cc.fill = FILL_BLOQ
        cc.border = BORDE
        cc.alignment = WRAP
    sig.row_dimensions[r].height = 34
    r += 1

r += 1
fila_hdr(sig, r, ["ORDEN", "CODIGO", "QUE HAY QUE HACER", "QUIEN", "CARRIL", "ESTADO", "DESBLOQUEA", "EN CADENA", "QUE FALTA PARA PODER EMPEZARLA"])


def bloque(fila, titulo, tareas, numerar):
    fila += 1
    sig.cell(row=fila, column=1, value=titulo)
    sig.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=9)
    for j in range(1, 10):
        cc = sig.cell(row=fila, column=j)
        cc.fill = FILL_FASE
        cc.font = F_FASE
        cc.border = BORDE
    for i, t in enumerate(tareas, start=1):
        fila += 1
        falta = ", ".join(f"{d} ({ETIQUETA[POR_CODIGO[d]['estado']].lower()}, {POR_CODIGO[d]['responsable']})" for d in t["falta"])
        valores = [
            i if numerar else "—",
            t["codigo"],
            t["tarea"],
            t["responsable"],
            t["carril"],
            ETIQUETA[t["estado"]],
            t["desbloquea"] or "",
            t["cadena"] or "",
            falta if falta else "nada: se puede empezar ya",
        ]
        for j, v in enumerate(valores, start=1):
            c = sig.cell(row=fila, column=j, value=v)
            c.border = BORDE
            c.font = F_TXT
            c.alignment = WRAP if j in (3, 4, 9) else CENTRO
            if j == 2:
                c.number_format = "@"
        if t["estado"] == "en_curso":
            sig.cell(row=fila, column=6).fill = FILL_CURSO
        elif t["estado"] == "bloqueada":
            sig.cell(row=fila, column=6).fill = FILL_BLOQ
        if t["es_ceo"]:
            sig.cell(row=fila, column=4).fill = FILL_CEO
        if t["carril"] == "MOTOR":
            sig.cell(row=fila, column=5).fill = FILL_NOTA
    return fila


r = bloque(r, f"SE PUEDE TRABAJAR YA — {len(COLA)} tareas, en este orden", COLA, True)
r = bloque(r, f"ESPERANDO A OTRA TAREA — {len(ESPERANDO)} tareas, en orden de codigo WBS", ESPERANDO, False)
r += 2
sig.cell(row=r, column=1, value=(
    "Aviso de la regla 8 (CLAUDE.md): si la cola de arriba solo tuviera tareas de MOTOR, hay que parar y avisar al CEO; "
    "significa que la cola esta mal llena. Ahora mismo la cola tiene "
    f"{sum(1 for t in COLA if t['carril'] == 'PRODUCTO')} de producto y {sum(1 for t in COLA if t['carril'] == 'MOTOR')} de motor."
))
sig.cell(row=r, column=1).font = F_TXT
sig.cell(row=r, column=1).alignment = WRAP
sig.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
sig.cell(row=r, column=1).fill = FILL_NOTA
sig.row_dimensions[r].height = 46
sig.freeze_panes = "A9"  # deja siempre a la vista los dos AHORA y la cabecera
anchos(sig, [8, 11, 56, 21, 11, 12, 12, 11, 42])

# ================================================================ 3. CALENDARIO
cal = wb.create_sheet("CALENDARIO")
cuerpo_cal = seccion("Calendario del mes")
objetivo = ""
for linea in cuerpo_cal.splitlines():
    if "Objetivo del mes" in linea:
        objetivo = limpiar(linea)
f = cabecera(cal, "CALENDARIO — 29 julio a 1 septiembre de 2026", 5, objetivo)
tabla_cal = tablas(cuerpo_cal)[0]
fila_hdr(cal, f, [limpiar(x) for x in tabla_cal[0]])
r = f
for filaf in tabla_cal[1:]:
    r += 1
    for j, v in enumerate(filaf, start=1):
        c = cal.cell(row=r, column=j, value=limpiar(v))
        c.border = BORDE
        c.font = F_TXT
        c.alignment = WRAP
    if limpiar(filaf[0]) == "GM":
        for j in range(1, 6):
            cal.cell(row=r, column=j).fill = FILL_CURSO
r += 2
no_cabe = [limpiar(l) for l in cuerpo_cal.splitlines() if l.startswith("**Lo que NO cabe")]
if no_cabe:
    cal.cell(row=r, column=1, value=no_cabe[0]).font = F_TXT
    cal.cell(row=r, column=1).alignment = WRAP
    cal.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cal.cell(row=r, column=1).fill = FILL_NOTA
    cal.row_dimensions[r].height = 46
anchos(cal, [9, 16, 46, 40, 30])

# ================================================================ 4. PUERTAS
pue = wb.create_sheet("PUERTAS")
f = cabecera(pue, "PUERTAS — donde decide el CEO", 2, "Criterios tal y como estan escritos en WBS.md (G1 corregida el 30/07/2026).")
fila_hdr(pue, f, ["PUERTA", "QUE SE DECIDE Y CON QUE CRITERIO"])
r = f
for linea in seccion("Puertas").splitlines():
    if not linea.strip().startswith("- **"):
        continue
    m = re.match(r"- \*\*(G[0-9M]+)\*\*(.*)", linea.strip())
    if not m:
        continue
    r += 1
    pue.cell(row=r, column=1, value=m.group(1)).font = F_BOLD
    pue.cell(row=r, column=2, value=limpiar(m.group(2)).lstrip(": "))
    for j in (1, 2):
        cc = pue.cell(row=r, column=j)
        cc.border = BORDE
        cc.alignment = WRAP
        if j == 2:
            cc.font = F_TXT
anchos(pue, [12, 118])

# ================================================================ 5. EQUIPO
eq = wb.create_sheet("EQUIPO")
cuerpo_eq = seccion("Equipo de agentes")
f = cabecera(eq, "EQUIPO DE AGENTES — un modelo por puesto", 5,
             "Regla 29: identificador exacto del modelo, nunca un alias, y ningun agente sin modelo.")
tabla_eq = tablas(cuerpo_eq)[0]
fila_hdr(eq, f, [limpiar(x) for x in tabla_eq[0]])
r = volcar(eq, f, [[limpiar(v) for v in filaf] for filaf in tabla_eq[1:]])
r += 2
for linea in cuerpo_eq.splitlines():
    if linea.startswith("**Precios") or linea.startswith("> Aviso"):
        eq.cell(row=r, column=1, value=limpiar(linea.lstrip("> ")))
        eq.cell(row=r, column=1).font = F_TXT
        eq.cell(row=r, column=1).alignment = WRAP
        eq.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        eq.cell(row=r, column=1).fill = FILL_NOTA
        eq.row_dimensions[r].height = 46
        r += 2
anchos(eq, [22, 50, 38, 16, 18])

# ================================================================ 6. REGLAS
# Tarea 03.01.16: se lee de CLAUDE.md, nunca de la tabla espejo del WBS (D-16 la derogo alli).
reg = wb.create_sheet("REGLAS")
f = cabecera(
    reg, f"LAS {len(REGLAS)} REGLAS — mandan sobre cualquier otra instruccion", 2,
    "Leidas de CLAUDE.md, seccion «## Las N reglas» (D-16: unica lista normativa). No se lee "
    "la tabla espejo que tenia el WBS: quedo derogada y vacia.",
)
fila_hdr(reg, f, ["Nº", "REGLA"])
r = f
for numero, texto_regla in REGLAS:
    r += 1
    reg.cell(row=r, column=1, value=numero).alignment = CENTRO
    reg.cell(row=r, column=2, value=recortar(texto_regla, 700, fuente="CLAUDE.md"))
    for j in (1, 2):
        cc = reg.cell(row=r, column=j)
        cc.border = BORDE
        cc.font = F_TXT
        if j == 2:
            cc.alignment = WRAP
anchos(reg, [6, 130])

# ================================================================ 7. DECISIONES
# Tarea 03.01.16: se lee de 00-direccion/DECISIONES.md, nunca de la tabla espejo del WBS
# (esa tabla se quedo congelada en D-21 y nunca llego a recoger D-22 a D-25).
dec = wb.create_sheet("DECISIONES")
f = cabecera(
    dec, f"REGISTRO DE DECISIONES ({len(DECISIONES)}) — solo se añade, nunca se reescribe (regla 21 de CLAUDE.md)", 4,
    "Leido de 00-direccion/DECISIONES.md, una fila por cada 'D-N'. No se lee la tabla espejo del WBS.",
)
fila_hdr(dec, f, ["ID", "FECHA", "DECISIÓN", "DETALLE"])
r = f
for d in DECISIONES:
    r += 1
    valores = [d["id"], d["fecha"], d["decision"], recortar(d["detalle"], 900, fuente="00-direccion/DECISIONES.md")]
    for j, v in enumerate(valores, start=1):
        c = dec.cell(row=r, column=j, value=v)
        c.border = BORDE
        c.font = F_TXT
        c.alignment = WRAP
        if j == 1:
            c.number_format = "@"
anchos(dec, [8, 12, 55, 75])

# ================================================================ 8. (hoja TRASPLANTE gb2 retirada)
# La hoja se construia de la seccion "Trasplante desde gb2" de WBS.md, eliminada el 03/08/2026
# por D-21 (el CEO ordena construir desde cero y no mirar gb2). Sin seccion no hay hoja: se
# retira aqui en vez de dejar el generador reventando con KeyError.

# ================================================================ 9. LECCIONES
# Tarea 03.01.16: se lee de 00-direccion/LECCIONES.md, nunca de la tabla espejo del WBS
# (esa tabla se quedo congelada en L-018 y nunca llego a recoger L-019 a L-031).
lec = wb.create_sheet("LECCIONES")
f = cabecera(
    lec, f"LECCIONES APRENDIDAS ({len(LECCIONES)})", 5,
    "Leidas de 00-direccion/LECCIONES.md, una fila por cada 'L-NNN'. No se lee la tabla espejo del WBS.",
)
fila_hdr(lec, f, ["ID", "SÍNTOMA", "CAUSA RAÍZ", "REGLA", "EVENTO"])
r = f
for l in LECCIONES:
    r += 1
    valores = [
        l["id"], l["sintoma"],
        recortar(l["causa"], 500, fuente="00-direccion/LECCIONES.md"),
        recortar(l["regla"], 400, fuente="00-direccion/LECCIONES.md"),
        recortar(l["evento"], 400, fuente="00-direccion/LECCIONES.md"),
    ]
    for j, v in enumerate(valores, start=1):
        c = lec.cell(row=r, column=j, value=v)
        c.border = BORDE
        c.font = F_TXT
        c.alignment = WRAP
        if j == 1:
            c.number_format = "@"
anchos(lec, [8, 50, 50, 42, 42])

# ================================================================ 10. AUTONOMIA
aut = wb.create_sheet("AUTONOMIA")
f = cabecera(aut, "AUTONOMIA — que llega al CEO y que no", 2)
fila_hdr(aut, f, ["CUANDO", "QUE PASA"])
r = f
for parrafo in seccion("Autonomía: qué llega al CEO").split("\n\n"):
    p = parrafo.strip()
    if not p.startswith("**"):
        continue
    m = re.match(r"\*\*(.+?):?\*\*:?\s*(.*)", p, flags=re.S)
    if not m:
        continue
    r += 1
    aut.cell(row=r, column=1, value=limpiar(m.group(1))).font = F_BOLD
    aut.cell(row=r, column=2, value=limpiar(m.group(2).replace(" · ", "\n· ")))
    for j in (1, 2):
        cc = aut.cell(row=r, column=j)
        cc.border = BORDE
        cc.alignment = WRAP
        if j == 2:
            cc.font = F_TXT
r += 2
aut.cell(row=r, column=1, value="FORMATO OBLIGATORIO DE FICHA DE DECISION (regla 10)").font = F_BOLD
r += 1
plantilla = seccion("Plantilla de ficha de decisión")
bloque = re.search(r"```(.*?)```", plantilla, flags=re.S)
if bloque:
    aut.cell(row=r, column=1, value=bloque.group(1).strip()).font = F_TXT
    aut.cell(row=r, column=1).alignment = WRAP
    aut.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    aut.cell(row=r, column=1).fill = FILL_NOTA
    aut.row_dimensions[r].height = 100
anchos(aut, [40, 96])

# ================================================================ 11. LIMITES
lim = wb.create_sheet("LIMITES CEO")
f = cabecera(lim, "LIMITES DEL CEO", 1)
r = f
for linea in seccion("Límites del CEO").splitlines():
    if linea.strip().startswith("- "):
        r += 1
        c = lim.cell(row=r, column=1, value=limpiar(linea.strip()[2:]))
        c.font = F_TXT
        c.alignment = WRAP
        c.border = BORDE
        lim.row_dimensions[r].height = 34
anchos(lim, [130])

wb.save(SALIDA)
n_tareas = sum(len(t) for _, _, t in FASES)
print(f"OK -> {SALIDA}")
print(f"   fases: {len(FASES)} · tareas: {n_tareas} · hojas: {len(wb.sheetnames)}")
for cod, nombre, tareas in FASES:
    cuenta = {e: sum(1 for t in tareas if t['estado'] == e) for e in ESTADOS}
    print(f"   {cod} {nombre[:38]:40s} {len(tareas):2d} tareas  {cuenta}")
if AVISOS:
    print("\nAVISOS de lectura del WBS (revisar a mano):")
    for a in AVISOS:
        print("   !", a)

# ---- que ha cambiado desde la generacion anterior (no depende de la memoria de nadie)
INSTANTANEA = Path(__file__).resolve().parent / "ultimo_estado.json"
ahora = {t["codigo"]: t["estado"] for t in TODAS}
print("\nCAMBIOS DESDE LA ULTIMA GENERACION:")
if INSTANTANEA.exists():
    previo = json.loads(INSTANTANEA.read_text(encoding="utf-8")).get("tareas", {})
    nuevas = [c for c in ahora if c not in previo]
    idas = [c for c in previo if c not in ahora]
    movidas = [(c, previo[c], ahora[c]) for c in ahora if c in previo and previo[c] != ahora[c]]
    if not (nuevas or idas or movidas):
        print("   nada: el WBS no ha cambiado de estado desde la ultima vez")
    for c in nuevas:
        print(f"   NUEVA      {c}  ({ahora[c]})  {POR_CODIGO[c]['tarea'][:60]}")
    for c, antes, despues in movidas:
        print(f"   CAMBIA     {c}  {antes} -> {despues}  {POR_CODIGO[c]['tarea'][:52]}")
    for c in idas:
        print(f"   DESAPARECE {c}  (estaba {previo[c]}) — ojo: los codigos no se borran (regla 3)")
else:
    print("   primera generacion con registro: no hay con que comparar")
INSTANTANEA.write_text(json.dumps({
    "_aviso": "Instantanea para calcular cambios entre generaciones. NO es fuente de verdad: lo es 00-direccion/WBS.md",
    "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
    "tareas": ahora,
}, indent=1, ensure_ascii=False), encoding="utf-8")

print("\nEstado leido de cada tarea (contrastar con WBS.md):")
for t in TODAS:
    print(f"   {t['codigo']}  {t['estado']:10s} {t['tarea'][:60]}")
